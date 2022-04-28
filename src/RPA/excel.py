"""_summary_
    Automation of Excel
"""
import openpyxl as xl

# Criar-uma planilha (book)
book = xl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Frutas')

# Como selecionar uma página
frutas_page = book['Frutas']
frutas_page.append(["Banana", '5', 'R$3,90'])
frutas_page.append(['Fruta 2', "2", 'R$15,90'])
frutas_page.append(['Fruta 3', '10', 'R$30,90'])
frutas_page.append(['Fruta 4', '2', 'R$50,50'])

# Salvar a planilha
book.save('Planilha de Compras.xlsx')


# Carregando arquivo
book = xl.load_workbook('Planilha de Compras.xlsx')
# Selecionando uma página
frutas_page = book['Frutas']
# Imprimindo os dados de cada linha
for rows in frutas_page.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        if cell.value == "Banana":
            cell.value = 'Fruta 1'
# Salvar as alterações
book. save('Planilha de Compras v2.xlsx')